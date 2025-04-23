import io
from datetime import datetime
from typing import List, Type
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import pandas as pd
from loguru import logger
from pydantic import BaseModel
from starlette.responses import StreamingResponse


async def export_excel(schema: Type[BaseModel], file_name: str, data_list: List[BaseModel] = []) -> StreamingResponse:
    """
    Export a template or data as an Excel file with Microsoft YaHei font for all cells and auto-width headers.
    """
    field_names = list(schema.model_fields.keys())
    user_export_df = pd.DataFrame(columns=field_names)
    if data_list:
        data_dicts = [item.model_dump() for item in data_list]
        user_export_df = pd.concat([user_export_df, pd.DataFrame(data_dicts)], ignore_index=True)

    filename = f"{file_name}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    stream = io.BytesIO()

    try:
        with pd.ExcelWriter(stream, engine="openpyxl") as writer:
            user_export_df.to_excel(writer, index=False, sheet_name=filename)
            worksheet = writer.sheets[filename]
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.font = Font(name="Microsoft YaHei")
            for idx, col in enumerate(user_export_df.columns, 1):
                column_letter = get_column_letter(idx)
                column_width = max(len(str(col)), 15)
                worksheet.column_dimensions[column_letter].width = column_width

        stream.seek(0)
        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as e:
        logger.error(f"Failed to export Excel: {e}")
        raise
